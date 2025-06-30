import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from collections import defaultdict

# --- 1. Definición de Entidades (Usando Clases para mejor estructura) ---

class Profesor:
    def __init__(self, nombre, materias_que_imparte):
        self.nombre = nombre
        self.materias_que_imparte = materias_que_imparte # Lista de nombres de materias

class Aula:
    def __init__(self, nombre, capacidad):
        self.nombre = nombre
        self.capacidad = capacidad

class Materia:
    def __init__(self, nombre, duracion_bloques, requiere_capacidad=0):
        self.nombre = nombre
        self.duracion_bloques = duracion_bloques # Cuántos bloques de tiempo dura
        self.requiere_capacidad = requiere_capacidad # Capacidad mínima de aula (ej. si es un grupo grande)

class ClaseRequerida:
    """Una instancia de una materia que necesita ser programada."""
    def __init__(self, id_clase, materia_obj, grupo, estudiantes=0):
        self.id = id_clase # Identificador único para esta instancia de clase (ej. "MATH101-A")
        self.materia = materia_obj # Objeto Materia
        self.grupo = grupo
        self.estudiantes = estudiantes # Número de estudiantes para verificar capacidad de aula
        self.asignada = False # Para rastrear si ya ha sido programada

# --- 2. Horario y Lógica de Asignación ---

class GeneradorHorarios:
    def __init__(self, profesores, aulas, materias, franjas_por_dia=5, dias_semana=["Lunes", "Martes", "Miércoles", "Jueves", "Viernes"]):
        self.profesores = profesores
        self.aulas = aulas
        self.materias = {m.nombre: m for m in materias} # Diccionario para acceso rápido
        self.franjas_por_dia = franjas_por_dia
        self.dias_semana = dias_semana

        # Estructura principal del horario:
        # horario[dia][franja_hora] = {"aula": Aula, "profesor": Profesor, "clase": ClaseRequerida}
        self.horario = defaultdict(lambda: defaultdict(dict))

        # Disponibilidad para rastrear conflictos
        # disponibilidad_profesores[dia][franja_hora] = [Profesor.nombre, ...]
        self.disponibilidad_profesores = defaultdict(lambda: defaultdict(list))
        # disponibilidad_aulas[dia][franja_hora] = [Aula.nombre, ...]
        self.disponibilidad_aulas = defaultdict(lambda: defaultdict(list))
        
        self.clases_no_asignadas = []

    def _es_franja_valida(self, dia_idx, franja_inicio, duracion):
        """Verifica si las franjas requeridas no exceden el límite diario."""
        return all(0 <= (franja_inicio + i) < self.franjas_por_dia for i in range(duracion))

    def _es_disponible(self, dia_idx, franja_inicio, duracion, profesor_nombre, aula_nombre):
        """Verifica si profesor y aula están disponibles para la duración requerida."""
        for i in range(duracion):
            franja_actual = franja_inicio + i
            if profesor_nombre in self.disponibilidad_profesores[self.dias_semana[dia_idx]][franja_actual]:
                return False # Profesor ya ocupado
            if aula_nombre in self.disponibilidad_aulas[self.dias_semana[dia_idx]][franja_actual]:
                return False # Aula ya ocupada
        return True

    def _asignar_bloque(self, dia_idx, franja_inicio, duracion, profesor_obj, aula_obj, clase_obj):
        """Asigna la clase y marca como ocupado."""
        for i in range(duracion):
            franja_actual = franja_inicio + i
            dia_nombre = self.dias_semana[dia_idx]
            
            self.horario[dia_nombre][franja_actual] = {
                "aula": aula_obj,
                "profesor": profesor_obj,
                "clase": clase_obj
            }
            self.disponibilidad_profesores[dia_nombre][franja_actual].append(profesor_obj.nombre)
            self.disponibilidad_aulas[dia_nombre][franja_actual].append(aula_obj.nombre)
        clase_obj.asignada = True # Marcar la clase como asignada
        print(f"  ASIGNADO: '{clase_obj.materia.nombre}' ({clase_obj.grupo}) con {profesor_obj.nombre} en {aula_obj.nombre} el {self.dias_semana[dia_idx]} de la franja {franja_inicio} a {franja_inicio + duracion - 1}")


    def generar(self, clases_a_programar):
        print("--- Iniciando Generación de Horario ---")
        clases_pendientes = sorted(clases_a_programar, key=lambda c: c.materia.duracion_bloques, reverse=True) # Priorizar clases más largas
        
        for clase_obj in clases_pendientes:
            if clase_obj.asignada:
                continue

            materia = clase_obj.materia
            duracion = materia.duracion_bloques
            print(f"Intentando programar: '{materia.nombre}' ({clase_obj.grupo})")

            clase_asignada_en_este_ciclo = False
            for dia_idx, dia_nombre in enumerate(self.dias_semana):
                if clase_asignada_en_este_ciclo: break # Si ya se asignó, pasa a la siguiente clase
                for franja_inicio in range(self.franjas_por_dia - duracion + 1): # - duracion + 1 para asegurar que cabe
                    if not self._es_franja_valida(dia_idx, franja_inicio, duracion):
                        continue

                    # Buscar profesor
                    for profesor in self.profesores:
                        if materia.nombre in profesor.materias_que_imparte:
                            # Buscar aula
                            for aula in self.aulas:
                                if aula.capacidad >= clase_obj.estudiantes: # Verifica capacidad
                                    if self._es_disponible(dia_idx, franja_inicio, duracion, profesor.nombre, aula.nombre):
                                        self._asignar_bloque(dia_idx, franja_inicio, duracion, profesor, aula, clase_obj)
                                        clase_asignada_en_este_ciclo = True
                                        break # Salir del bucle de aulas
                            if clase_asignada_en_este_ciclo:
                                break # Salir del bucle de profesores
                    if clase_asignada_en_este_ciclo:
                        break # Salir del bucle de franjas

            if not clase_asignada_en_este_ciclo:
                self.clases_no_asignadas.append(clase_obj)
                print(f"  NO SE PUDO ASIGNAR: '{materia.nombre}' ({clase_obj.grupo})")

        print("--- Generación de Horario Finalizada ---")
        if self.clases_no_asignadas:
            print("\nClases no asignadas (posibles conflictos o falta de recursos):")
            for clase in self.clases_no_asignadas:
                print(f"- {clase.materia.nombre} ({clase.grupo})")

    def mostrar_horario_consola(self):
        """Imprime el horario en la consola de forma rudimentaria."""
        print("\n--- Horario Generado (Consola) ---")
        franjas_str = [f"{i:02d}:00-{i+1:02d}:00" for i in range(self.franjas_por_dia)]

        for dia_nombre in self.dias_semana:
            print(f"\n--- {dia_nombre.upper()} ---")
            for franja_idx, franja_str in enumerate(franjas_str):
                clase_info = self.horario[dia_nombre].get(franja_idx)
                if clase_info:
                    print(f"  {franja_str}: Aula '{clase_info['aula'].nombre}' | Prof '{clase_info['profesor'].nombre}' | Clase '{clase_info['clase'].materia.nombre}' ({clase_info['clase'].grupo})")
                else:
                    print(f"  {franja_str}: Disponible")

    def exportar_a_excel(self, nombre_archivo="horario_generado.xlsx"):
        """Exporta el horario a un archivo Excel."""
        print(f"\n--- Exportando Horario a '{nombre_archivo}' ---")
        wb = openpyxl.Workbook()
        
        # Estilos
        bold_font = Font(bold=True, size=12)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid") # Gris claro

        # Crear una hoja por cada aula
        for aula in self.aulas:
            ws = wb.create_sheet(title=f"Aula {aula.nombre}")
            
            # Encabezados de días
            ws.cell(row=1, column=1, value="Hora").font = bold_font
            ws.cell(row=1, column=1).alignment = center_align
            ws.cell(row=1, column=1).fill = header_fill

            for col_idx, dia_nombre in enumerate(self.dias_semana):
                cell = ws.cell(row=1, column=col_idx + 2, value=dia_nombre)
                cell.font = bold_font
                cell.alignment = center_align
                cell.fill = header_fill
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx + 2)].width = 25 # Ancho de columna

            # Horas (franjas)
            for row_idx, franja_idx in enumerate(range(self.franjas_por_dia)):
                franja_str = f"{franja_idx:02d}:00-{franja_idx+1:02d}:00"
                cell = ws.cell(row=row_idx + 2, column=1, value=franja_str)
                cell.font = bold_font
                cell.alignment = center_align
                cell.border = thin_border
                ws.row_dimensions[row_idx + 2].height = 50 # Alto de fila

            # Llenar el horario del aula
            for dia_idx, dia_nombre in enumerate(self.dias_semana):
                for franja_idx in range(self.franjas_por_dia):
                    clase_en_esta_celda = None
                    # Buscar si hay una clase en esta aula, en este día y franja
                    if franja_idx in self.horario[dia_nombre]:
                        detalle_clase = self.horario[dia_nombre][franja_idx]
                        if detalle_clase["aula"].nombre == aula.nombre:
                            clase_en_esta_celda = detalle_clase

                    cell = ws.cell(row=franja_idx + 2, column=dia_idx + 2)
                    if clase_en_esta_celda:
                        clase_obj = clase_en_esta_celda['clase']
                        profesor_obj = clase_en_esta_celda['profesor']
                        cell.value = (f"{clase_obj.materia.nombre}\n"
                                      f"Grupo: {clase_obj.grupo}\n"
                                      f"Prof: {profesor_obj.nombre}")
                        cell.fill = PatternFill(start_color="E0FFFF", end_color="E0FFFF", fill_type="solid") # Azul claro
                    else:
                        cell.value = "Disponible"
                        cell.fill = PatternFill(start_color="F0FFF0", end_color="F0FFF0", fill_type="solid") # Verde claro

                    cell.alignment = center_align
                    cell.border = thin_border

        # Crear una hoja para resumen por profesor
        ws_profesores = wb.create_sheet(title="Horario Profesores")
        ws_profesores.cell(row=1, column=1, value="Profesor").font = bold_font
        ws_profesores.cell(row=1, column=1).alignment = center_align
        ws_profesores.cell(row=1, column=1).fill = header_fill

        for col_idx, dia_nombre in enumerate(self.dias_semana):
            cell = ws_profesores.cell(row=1, column=col_idx + 2, value=dia_nombre)
            cell.font = bold_font
            cell.alignment = center_align
            cell.fill = header_fill
            ws_profesores.column_dimensions[openpyxl.utils.get_column_letter(col_idx + 2)].width = 25

        for row_idx, franja_idx in enumerate(range(self.franjas_por_dia)):
            franja_str = f"{franja_idx:02d}:00-{franja_idx+1:02d}:00"
            cell = ws_profesores.cell(row=row_idx + 2, column=1, value=franja_str)
            cell.font = bold_font
            cell.alignment = center_align
            cell.border = thin_border
            ws_profesores.row_dimensions[row_idx + 2].height = 50

        # Llenar el horario de profesores
        for profesor_obj in self.profesores:
            row_start = ws_profesores.max_row + 1 if ws_profesores.max_row > 1 else 2 # Empieza después de encabezados

            # Fila para el nombre del profesor
            profesor_name_cell = ws_profesores.cell(row=row_start, column=1, value=profesor_obj.nombre)
            profesor_name_cell.font = bold_font
            profesor_name_cell.alignment = center_align
            profesor_name_cell.fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid") # Naranja claro
            profesor_name_cell.border = thin_border

            for dia_idx, dia_nombre in enumerate(self.dias_semana):
                for franja_idx in range(self.franjas_por_dia):
                    clase_en_esta_celda = None
                    if franja_idx in self.horario[dia_nombre]:
                        detalle_clase = self.horario[dia_nombre][franja_idx]
                        if detalle_clase["profesor"].nombre == profesor_obj.nombre:
                            clase_en_esta_celda = detalle_clase
                    
                    cell = ws_profesores.cell(row=franja_idx + row_start, column=dia_idx + 2)
                    if clase_en_esta_celda:
                        clase_obj = clase_en_esta_celda['clase']
                        aula_obj = clase_en_esta_celda['aula']
                        cell.value = (f"{clase_obj.materia.nombre}\n"
                                      f"Grupo: {clase_obj.grupo}\n"
                                      f"Aula: {aula_obj.nombre}")
                        cell.fill = PatternFill(start_color="E0FFFF", end_color="E0FFFF", fill_type="solid")
                    else:
                        cell.value = "Disponible"
                        cell.fill = PatternFill(start_color="F0FFF0", end_color="F0FFF0", fill_type="solid")

                    cell.alignment = center_align
                    cell.border = thin_border
            # Fusionar la celda del nombre del profesor si hay múltiples franjas
            if self.franjas_por_dia > 1:
                ws_profesores.merge_cells(start_row=row_start, end_row=row_start + self.franjas_por_dia -1 , start_column=1, end_column=1)


        # Eliminar la hoja por defecto creada al inicio
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        try:
            wb.save(nombre_archivo)
            print(f"Horario exportado exitosamente a '{nombre_archivo}'")
        except Exception as e:
            print(f"Error al guardar el archivo Excel: {e}")

# --- 3. Datos de Prueba (Puedes modificar esto) ---

# Profesores
profesores_data = [
    Profesor("Dr. López", ["Matemáticas I", "Matemáticas II", "Álgebra"]),
    Profesor("Mtra. García", ["Literatura", "Español", "Historia"]),
    Profesor("Ing. Pérez", ["Física I", "Química", "Programación"]),
    Profesor("Lic. Díaz", ["Educación Física", "Arte"])
]

# Aulas
aulas_data = [
    Aula("A101", 30),
    Aula("A102", 25),
    Aula("B201", 40),
    Aula("Lab Quimica", 20),
    Aula("Gimnasio", 100)
]

# Materias
materias_data = [
    Materia("Matemáticas I", 2),       # Dura 2 bloques
    Materia("Literatura", 1),          # Dura 1 bloque
    Materia("Física I", 2),
    Materia("Historia", 1),
    Materia("Programación", 3, requiere_capacidad=25), # Dura 3 bloques, requiere aula de 25+
    Materia("Educación Física", 2, requiere_capacidad=50) # Requiere aula grande
]

# Clases que necesitamos programar
clases_requeridas = [
    ClaseRequerida("MI1A", materias_data[0], "1A", estudiantes=25),
    ClaseRequerida("LIT1B", materias_data[1], "1B", estudiantes=20),
    ClaseRequerida("FI1A", materias_data[2], "1A", estudiantes=22),
    ClaseRequerida("HIST2A", materias_data[3], "2A", estudiantes=35), # Este grupo es grande
    ClaseRequerida("PROG3C", materias_data[4], "3C", estudiantes=28),
    ClaseRequerida("EF4D", materias_data[5], "4D", estudiantes=60), # Este grupo es muy grande para casi todas las aulas
    ClaseRequerida("MI1C", materias_data[0], "1C", estudiantes=20),
    ClaseRequerida("LIT1C", materias_data[1], "1C", estudiantes=20),
]

# --- 4. Ejecución del Generador ---
if __name__ == "__main__":
    # Puedes ajustar el número de franjas por día y los días de la semana
    generador = GeneradorHorarios(
        profesores_data,
        aulas_data,
        materias_data,
        franjas_por_dia=5,
        dias_semana=["Lunes", "Martes", "Miércoles", "Jueves", "Viernes"]
    )

    generador.generar(clases_requeridas)
    generador.mostrar_horario_consola()
    generador.exportar_a_excel("horario_escolar.xlsx")